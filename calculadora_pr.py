import customtkinter as ctk
from datetime import datetime
import pytz
import math
from tkinter import filedialog
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font

class TransactionType:
    SALE = 'Compraventa'
    MORTGAGE = 'Hipoteca Nueva'
    MORTGAGE_CANCELLATION = 'Cancelación de Hipoteca'
    DONATION = 'Donación'
    OTHER = 'Otro (con cuantía)'

    @classmethod
    def values(cls):
        return [cls.SALE, cls.MORTGAGE, cls.MORTGAGE_CANCELLATION, cls.DONATION, cls.OTHER]

def calculate_fees(value: float, transaction_type: str, is_social_interest: bool, notary_percentage: float, num_copies: int) -> dict:
    notary_fee = 0
    if is_social_interest and transaction_type != TransactionType.MORTGAGE_CANCELLATION:
        fee_from_percent = value * (notary_percentage / 100)
        min_fee_from_percent = value * 0.0025
        notary_fee = max(fee_from_percent, min_fee_from_percent, 250)
    elif transaction_type == TransactionType.MORTGAGE_CANCELLATION:
        fee_from_percent = value * (notary_percentage / 100)
        min_fee_from_percent = value * 0.005
        notary_fee = max(fee_from_percent, min_fee_from_percent, 250)
    else:
        if value <= 10000:
            notary_fee = 150
        else:
            percentage_to_apply = max(0.5, min(1.0, notary_percentage))
            fee_from_percent = value * (percentage_to_apply / 100)
            notary_fee = max(fee_from_percent, 250)

    registry_fee = 0
    if value > 0:
        if value <= 1000:
            registry_fee = 2.00
        elif value <= 25000:
            registry_fee = math.ceil(value / 1000) * 2.00
        else:
            additional_thousands = math.ceil((value - 25000) / 1000)
            registry_fee = 50.00 + (additional_thousands * 4.00)

    ri_orig, ri_copias_total = 0, 0
    if value > 0:
        if value <= 250:
            ri_orig, ri_copias_total = 0.50, 0.20 * num_copies
        elif value <= 500:
            ri_orig, ri_copias_total = 1.00, 0.50 * num_copies
        elif value <= 1000:
            ri_orig, ri_copias_total = 2.00, 1.00 * num_copies
        elif value <= 5000:
            additional_thousands = math.ceil((value - 1000) / 1000)
            ri_orig = 2.00 + additional_thousands * 0.50
            ri_copias_total = (1.00 + additional_thousands * 0.20) * num_copies
        else:
            additional_thousands = math.ceil((value - 1000) / 1000)
            ri_orig = 2.00 + additional_thousands * 1.00
            ri_copias_total = (1.00 + additional_thousands * 0.50) * num_copies

    al_orig, al_copias_total = 0, 0
    applicable_types = [TransactionType.SALE, TransactionType.MORTGAGE, TransactionType.MORTGAGE_CANCELLATION]
    if transaction_type in applicable_types and value >= 25000:
        num_blocks = math.ceil(value / 50000)
        al_orig = num_blocks * 5.00
        al_copias_total = (num_blocks * 2.50) * num_copies

    presentation_fee = 15.00
    political_code_fee = 0.50
    if transaction_type == TransactionType.MORTGAGE_CANCELLATION:
        political_code_fee = 0

    total = (notary_fee + registry_fee + ri_orig + ri_copias_total +
             al_orig + al_copias_total + presentation_fee + political_code_fee)

    return {
        'Honorarios del Notario': notary_fee,
        'Arancel de Inscripción (Registro)': registry_fee,
        'Asiento de Presentación': presentation_fee,
        'Comprobante Código Político': political_code_fee,
        'ri_orig': ri_orig,
        'ri_copias': ri_copias_total,
        'al_orig': al_orig,
        'al_copias': al_copias_total,
        'total': total
    }

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Calculadora de Gastos Legales PR")
        self.geometry("900x780")
        self.grid_columnconfigure((0, 1), weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.history = []
        self.last_results, self.last_inputs = None, None
        self.transaction_value_var = ctk.StringVar(value="150000")
        self.transaction_type_var = ctk.StringVar(value=TransactionType.SALE)
        self.social_interest_var = ctk.BooleanVar()
        self.notary_percentage_var = ctk.StringVar(value="0.75") # Cambiado a StringVar
        self.num_copies_var = ctk.StringVar(value="1")

        self.create_widgets()
        self.perform_calculation()

    def create_widgets(self):
        self.left_panel = ctk.CTkFrame(self, fg_color="transparent")
        self.left_panel.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        self.left_panel.grid_columnconfigure(0, weight=1)
        self.input_frame = ctk.CTkFrame(self.left_panel)
        self.input_frame.grid(row=0, column=0, sticky="ew", pady=(0, 15))
        self.input_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(self.input_frame, text="Configuración del Cálculo", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, columnspan=2, pady=10)
        self.create_form_inputs(self.input_frame)
        ctk.CTkButton(self.input_frame, text="Calcular", command=self.perform_calculation).grid(row=6, column=0, columnspan=2, pady=15, padx=20, sticky="ew")
        self.results_frame = ctk.CTkFrame(self.left_panel, fg_color="transparent")
        self.results_frame.grid(row=1, column=0, sticky="nsew")
        self.results_frame.grid_columnconfigure(0, weight=1)
        self.right_panel = ctk.CTkFrame(self)
        self.right_panel.grid(row=0, column=1, padx=(0, 20), pady=20, sticky="nsew")
        self.right_panel.grid_rowconfigure(1, weight=1)
        history_header_frame = ctk.CTkFrame(self.right_panel, fg_color="transparent")
        history_header_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        history_header_frame.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(history_header_frame, text="Historial", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, sticky="w")
        ctk.CTkButton(history_header_frame, text="Limpiar", width=60, command=self.clear_history).grid(row=0, column=1, sticky="e")
        self.history_scroll_frame = ctk.CTkScrollableFrame(self.right_panel, label_text="", fg_color="transparent")
        self.history_scroll_frame.grid(row=1, column=0, sticky="nsew", padx=5)

    def create_form_inputs(self, parent):
        ctk.CTkLabel(parent, text="Valor de la Transacción ($)").grid(row=1, column=0, padx=20, pady=(5,10), sticky="w")
        ctk.CTkEntry(parent, textvariable=self.transaction_value_var).grid(row=1, column=1, padx=20, pady=(5,10), sticky="ew")
        ctk.CTkLabel(parent, text="Tipo de Transacción").grid(row=2, column=0, padx=20, pady=10, sticky="w")
        ctk.CTkComboBox(parent, variable=self.transaction_type_var, values=TransactionType.values()).grid(row=2, column=1, padx=20, pady=10, sticky="ew")
        ctk.CTkLabel(parent, text="No. de Copias Certificadas").grid(row=3, column=0, padx=20, pady=10, sticky="w")
        ctk.CTkEntry(parent, textvariable=self.num_copies_var, width=120).grid(row=3, column=1, padx=20, pady=10, sticky="w")
        ctk.CTkCheckBox(parent, text="¿Es Vivienda de Interés Social?", variable=self.social_interest_var).grid(row=4, column=0, columnspan=2, padx=20, pady=10, sticky="w")
        
        # --- CAMBIO: Reemplazo de Slider por Campo de Texto ---
        ctk.CTkLabel(parent, text="Honorarios Notariales (%)").grid(row=5, column=0, padx=20, pady=10, sticky="w")
        entry_frame = ctk.CTkFrame(parent, fg_color="transparent")
        entry_frame.grid(row=5, column=1, padx=20, pady=10, sticky="ew")
        ctk.CTkEntry(entry_frame, textvariable=self.notary_percentage_var).pack(side="left", fill="x", expand=True)
        ctk.CTkLabel(entry_frame, text="%").pack(side="left", padx=(5,0))

    def perform_calculation(self):
        try: value = float(self.transaction_value_var.get())
        except ValueError: value = 0.0
        try: num_copies = int(self.num_copies_var.get())
        except ValueError: num_copies = 0
        try: notary_percentage = float(self.notary_percentage_var.get())
        except ValueError: notary_percentage = 0.0

        self.last_inputs = {'value': value, 'transaction_type': self.transaction_type_var.get(), 'is_social_interest': self.social_interest_var.get(), 'notary_percentage': notary_percentage, 'num_copies': num_copies}
        self.last_results = calculate_fees(**self.last_inputs)
        self.display_results(self.last_results)
        if value > 0: self.add_to_history(self.last_inputs, self.last_results)

    def display_results(self, results: dict):
        for widget in self.results_frame.winfo_children(): widget.destroy()
        total_frame_card = ctk.CTkFrame(self.results_frame, fg_color=("gray85", "gray20"))
        total_frame_card.pack(fill="x", padx=0, pady=(0, 15))
        ctk.CTkLabel(total_frame_card, text="Gasto Total Estimado", font=ctk.CTkFont(size=14)).pack(pady=(10, 0))
        ctk.CTkLabel(total_frame_card, text=f"${results['total']:,.2f}", font=ctk.CTkFont(size=28, weight="bold"), text_color="#2CC985").pack(pady=(0, 10))
        fees_card = ctk.CTkFrame(self.results_frame)
        fees_card.pack(fill="x", expand=True, padx=0, pady=(0,15))
        ctk.CTkLabel(fees_card, text="Honorarios y Aranceles", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=10, pady=5)
        for key in ['Honorarios del Notario', 'Arancel de Inscripción (Registro)', 'Asiento de Presentación', 'Comprobante Código Político']:
            if results[key] > 0: self.create_result_row(fees_card, key, results[key])
        stamps_card = ctk.CTkFrame(self.results_frame)
        stamps_card.pack(fill="x", expand=True, padx=0, pady=0)
        ctk.CTkLabel(stamps_card, text="Sellos y Comprobantes", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=10, pady=5)
        if results['ri_orig'] > 0: self.create_result_row(stamps_card, "Sellos Rentas Internas (Original)", results['ri_orig'])
        if results['ri_copias'] > 0: self.create_result_row(stamps_card, f"Sellos Rentas Internas ({self.num_copies_var.get()} Copias)", results['ri_copias'])
        if results['al_orig'] > 0: self.create_result_row(stamps_card, "Sello Asistencia Legal (Original)", results['al_orig'])
        if results['al_copias'] > 0: self.create_result_row(stamps_card, f"Sello Asistencia Legal ({self.num_copies_var.get()} Copias)", results['al_copias'])
        export_frame = ctk.CTkFrame(self.results_frame, fg_color="transparent")
        export_frame.pack(fill="x", pady=(15,0))
        export_frame.grid_columnconfigure((0,1), weight=1)
        ctk.CTkButton(export_frame, text="Exportar a PDF", command=self.export_to_pdf).grid(row=0, column=0, padx=(0,5), sticky="ew")
        ctk.CTkButton(export_frame, text="Exportar a Excel", command=self.export_to_excel).grid(row=0, column=1, padx=(5,0), sticky="ew")

    def create_result_row(self, parent, label, value):
        row_frame = ctk.CTkFrame(parent, fg_color="transparent")
        row_frame.pack(fill="x", padx=10, pady=2)
        row_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(row_frame, text=label, wraplength=200, justify="left").grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(row_frame, text=f"${value:,.2f}", font=ctk.CTkFont(weight="bold")).grid(row=0, column=1, sticky="e")

    def add_to_history(self, inputs, results):
        tz = pytz.timezone('America/Puerto_Rico')
        timestamp = datetime.now(tz).strftime('%d/%m/%Y %I:%M %p')
        entry = {"inputs": inputs, "results": results, "timestamp": timestamp}
        self.history.insert(0, entry)
        self.history = self.history[:50]
        self.update_history_display()

    def update_history_display(self):
        for widget in self.history_scroll_frame.winfo_children(): widget.destroy()
        for entry in self.history:
            card = ctk.CTkButton(self.history_scroll_frame, fg_color=("gray85", "gray20"), hover_color=("gray80", "gray25"), text=f"{entry['inputs']['transaction_type']} - ${entry['inputs']['value']:,.2f}\n{entry['timestamp']}", anchor="w", command=lambda e=entry: self.load_from_history(e))
            card.pack(fill="x", padx=5, pady=5)

    def load_from_history(self, entry: dict):
        inputs = entry["inputs"]
        self.transaction_value_var.set(str(inputs["value"]))
        self.transaction_type_var.set(inputs["transaction_type"])
        self.social_interest_var.set(inputs["is_social_interest"])
        self.notary_percentage_var.set(str(inputs["notary_percentage"]))
        self.num_copies_var.set(str(inputs["num_copies"]))
        self.perform_calculation()

    def clear_history(self):
        self.history, self.last_results, self.last_inputs = [], None, None
        self.update_history_display()
        for widget in self.results_frame.winfo_children(): widget.destroy()

    def export_to_pdf(self):
        if not self.last_results: return
        filepath = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF Documents", "*.pdf")])
        if not filepath: return
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", size=12)
        pdf.set_font("Helvetica", 'B', 16)
        pdf.cell(0, 10, "Informe de Gastos Legales Estimados", 0, 1, 'C')
        pdf.ln(5)
        pdf.set_font("Helvetica", 'B', 12)
        pdf.cell(0, 10, "Parametros del Calculo", 0, 1, 'L')
        pdf.set_font("Helvetica", '', 12)
        pdf.cell(60, 8, "Valor de la Transaccion:", 0, 0)
        pdf.cell(0, 8, f"${self.last_inputs['value']:,.2f}", 0, 1)
        pdf.cell(60, 8, "Tipo de Transaccion:", 0, 0)
        pdf.cell(0, 8, self.last_inputs['transaction_type'], 0, 1)
        pdf.cell(60, 8, "Numero de Copias:", 0, 0)
        pdf.cell(0, 8, str(self.last_inputs['num_copies']), 0, 1)
        pdf.ln(8)
        def write_row(label, value):
            pdf.set_font("Helvetica", '', 12)
            pdf.cell(130, 8, label, 0, 0)
            pdf.set_font("Helvetica", 'B', 12)
            pdf.cell(0, 8, f"${value:,.2f}", 0, 1, 'R')
        pdf.set_font("Helvetica", 'B', 12)
        pdf.cell(0, 10, "Desglose de Gastos", "B", 1, 'L')
        pdf.ln(2)
        write_row("Honorarios del Notario", self.last_results['Honorarios del Notario'])
        write_row("Arancel de Inscripcion (Registro)", self.last_results['Arancel de Inscripción (Registro)'])
        write_row("Asiento de Presentacion", self.last_results['Asiento de Presentación'])
        write_row("Comprobante Codigo Politico", self.last_results['Comprobante Código Político'])
        pdf.ln(4)
        write_row("Sellos Rentas Internas (Original)", self.last_results['ri_orig'])
        write_row(f"Sellos Rentas Internas ({self.last_inputs['num_copies']} Copias)", self.last_results['ri_copias'])
        write_row("Sello Asistencia Legal (Original)", self.last_results['al_orig'])
        write_row(f"Sello Asistencia Legal ({self.last_inputs['num_copies']} Copias)", self.last_results['al_copias'])
        pdf.ln(4)
        pdf.set_font("Helvetica", 'B', 14)
        pdf.cell(130, 10, "Gasto Total Estimado", "T", 0)
        pdf.cell(0, 10, f"${self.last_results['total']:,.2f}", "T", 1, 'R')
        pdf.output(filepath)

    def export_to_excel(self):
        if not self.last_results: return
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not filepath: return
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Estimado de Gastos"
        bold_font = Font(bold=True)
        sheet.append(["Parametros del Calculo"])
        sheet["A1"].font = bold_font
        sheet.append(["Valor de la Transaccion:", f"${self.last_inputs['value']:,.2f}"])
        sheet.append(["Tipo de Transaccion:", self.last_inputs['transaction_type']])
        sheet.append(["Numero de Copias:", self.last_inputs['num_copies']])
        sheet.append([])
        sheet.append(["Concepto", "Monto"])
        sheet["A6"].font = bold_font
        sheet["B6"].font = bold_font
        rows = [
            ("Honorarios del Notario", self.last_results['Honorarios del Notario']),
            ("Arancel de Inscripcion (Registro)", self.last_results['Arancel de Inscripción (Registro)']),
            ("Asiento de Presentacion", self.last_results['Asiento de Presentación']),
            ("Comprobante Codigo Politico", self.last_results['Comprobante Código Político']),
            ("Sellos Rentas Internas (Original)", self.last_results['ri_orig']),
            (f"Sellos Rentas Internas ({self.last_inputs['num_copies']} Copias)", self.last_results['ri_copias']),
            ("Sello Asistencia Legal (Original)", self.last_results['al_orig']),
            (f"Sello Asistencia Legal ({self.last_inputs['num_copies']} Copias)", self.last_results['al_copias'])
        ]
        for row in rows:
            if row[1] > 0: sheet.append(row)
        sheet.append([])
        total_row_idx = sheet.max_row + 1
        sheet.cell(row=total_row_idx, column=1, value="Gasto Total Estimado").font = bold_font
        sheet.cell(row=total_row_idx, column=2, value=f"${self.last_results['total']:,.2f}").font = bold_font
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 20
        workbook.save(filepath)

if __name__ == "__main__":
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")
    app = App()
    app.mainloop()
